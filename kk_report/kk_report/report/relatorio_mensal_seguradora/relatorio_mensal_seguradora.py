from __future__ import unicode_literals

import frappe
from frappe import _
from frappe.utils import flt, getdate


def execute(filters=None):
    filters = frappe._dict(filters or {})
    columns = get_columns()
    data = get_data(filters)
    return columns, data


def get_columns():
    return [
        {
            "fieldname": "posting_date",
            "label": _("Data"),
            "fieldtype": "Data",
            "width": 110,
        },
        {
            "fieldname": "name",
            "label": _("Nº Factura"),
            "fieldtype": "Link",
            "options": "Sales Invoice",
            "width": 170,
        },
        {
            "fieldname": "membro_apolice",
            "label": _("Nº Membro"),
            "fieldtype": "Data",
            "width": 150,
        },
        {
            "fieldname": "net_total",
            "label": _("Valor (MZN)"),
            "fieldtype": "Currency",
            "width": 150,
        },
        {
            "fieldname": "total_taxes_and_charges",
            "label": _("IVA (MZN)"),
            "fieldtype": "Currency",
            "width": 150,
        },
        {
            "fieldname": "grand_total",
            "label": _("Total (MZN)"),
            "fieldtype": "Currency",
            "width": 150,
        },
    ]


def get_data(filters):
    conditions, values = get_conditions(filters)

    rows = frappe.db.sql(
        """
        SELECT
            si.posting_date,
            si.name,
            si.membro_apolice,
            si.net_total,
            si.total_taxes_and_charges,
            si.grand_total
        FROM
            `tabSales Invoice` si
        WHERE
            si.docstatus = 1
            {conditions}
        ORDER BY
            si.posting_date ASC,
            si.name ASC
        """.format(
            conditions=conditions
        ),
        values,
        as_dict=1,
    )

    data = []
    total_net = 0.0
    total_tax = 0.0
    total_grand = 0.0

    for row in rows:
        # Format date as dd/mm/yyyy — preserved as-is in all export formats
        if row.posting_date:
            row.posting_date = getdate(row.posting_date).strftime("%d/%m/%Y")

        total_net += flt(row.net_total)
        total_tax += flt(row.total_taxes_and_charges)
        total_grand += flt(row.grand_total)

        data.append(row)

    # Totals row — bold: 1 renders it bold in the web UI
    if data:
        data.append(
            {
                "posting_date": _("Total"),
                "name": "",
                "membro_apolice": "",
                "net_total": total_net,
                "total_taxes_and_charges": total_tax,
                "grand_total": total_grand,
                "bold": 1,
            }
        )

    return data


@frappe.whitelist()
def download_xlsx(filters=None):
    """Return a formatted Excel file with currency columns styled as #,##0.00."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if isinstance(filters, str):
        filters = frappe.parse_json(filters)
    filters = frappe._dict(filters or {})

    columns = get_columns()
    data = get_data(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio Mensal Seguradora"

    # Header row
    ws.append([col["label"] for col in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Currency column indices (1-based)
    currency_cols = [i + 1 for i, col in enumerate(columns) if col["fieldtype"] == "Currency"]

    # Data rows
    for row in data:
        row_values = [row.get(col["fieldname"], "") for col in columns]
        ws.append(row_values)

        current_row = ws.max_row
        is_total = bool(row.get("bold"))

        for col_idx in currency_cols:
            cell = ws.cell(row=current_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            if is_total:
                cell.font = Font(bold=True)

        if is_total:
            ws.cell(row=current_row, column=1).font = Font(bold=True)

    # Column widths
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(col.get("width", 120) // 7, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "relatorio_mensal_seguradora.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"


def get_conditions(filters):
    conditions = []
    values = frappe._dict()

    if filters.get("company"):
        conditions.append("si.company = %(company)s")
        values.company = filters.company

    if filters.get("date_from"):
        conditions.append("si.posting_date >= %(date_from)s")
        values.date_from = filters.date_from

    if filters.get("date_to"):
        conditions.append("si.posting_date <= %(date_to)s")
        values.date_to = filters.date_to

    if filters.get("seguradora"):
        conditions.append("si.seguradora = %(seguradora)s")
        values.seguradora = filters.seguradora

    if filters.get("status"):
        conditions.append("si.status = %(status)s")
        values.status = filters.status

    cond_str = ("AND " + " AND ".join(conditions)) if conditions else ""
    return cond_str, values
